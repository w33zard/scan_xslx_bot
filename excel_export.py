"""
Экспорт данных паспортов в Excel
"""
import os
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from ocr_extractor import EXCEL_COLUMNS

# Маппинг полей PassportResult (passport_ocr) → Excel
_PASSPORT_RESULT_TO_EXCEL = {
    "surname": "Фамилия",
    "name": "Имя",
    "patronymic": "Отчество",
    "birth_date": "Дата рождения",
    "birth_place": "Место рождения",
    "passport_series": None,  # объединяем с passport_number
    "passport_number": None,
    "issue_date": "Дата выдачи",
    "issue_place": "Кем выдан",
    "registration_address": "Адрес регистрации",
    "authority_code": None,
}


def _normalize_row(row: dict) -> dict:
    """
    Привести строку к плоскому формату Excel.
    Если пришёл PassportResult (doc_type, fields) — преобразовать.
    """
    if not isinstance(row, dict):
        return {}
    if "doc_type" not in row or "fields" not in row:
        return row

    fields = row.get("fields") or {}
    flat = {"№ п/п": row.get("№ п/п") or ""}

    def _val(key: str) -> str:
        f = fields.get(key)
        if f is None:
            return ""
        v = f.get("value") if isinstance(f, dict) else getattr(f, "value", None)
        return (v or "").strip()

    for fkey, col in _PASSPORT_RESULT_TO_EXCEL.items():
        if col:
            flat[col] = _val(fkey)

    s4, n6 = _val("passport_series"), _val("passport_number")
    if s4 and n6:
        flat["Серия и номер паспорта"] = f"{s4} {n6}".strip()
    elif s4 or n6:
        flat["Серия и номер паспорта"] = (s4 or n6).strip()

    for col in EXCEL_COLUMNS:
        if col not in flat:
            flat[col] = row.get(col, "")
    return flat


def normalize_results(data: list[dict]) -> list[dict]:
    """Привести список строк к плоскому формату Excel (если был PassportResult)."""
    if not data:
        return []
    out = []
    for i, r in enumerate(data, 1):
        row = _normalize_row(r)
        if not row.get("№ п/п"):
            row["№ п/п"] = str(i)
        out.append(row)
    return out


def _get_columns_from_template(template_path: str) -> list[str] | None:
    """Загрузить заголовки колонок из эталонного Excel"""
    if not os.path.exists(template_path):
        return None
    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]
        wb.close()
        return headers if headers else None
    except Exception:
        return None


def create_excel(
    data: list[dict],
    output_path: str,
    template_excel: str | None = None,
) -> str:
    """
    Создать Excel-файл с данными паспортов в формате как в эталоне
    """
    data = normalize_results(data)
    columns = _get_columns_from_template(template_excel) if template_excel else None
    columns = columns or EXCEL_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "Паспорта"

    # Заголовки
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Данные (data уже нормализован через normalize_results)
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Автоширина колонок — Кем выдан и Адрес: шире, чтобы текст был виден
    wide_cols = ("Кем выдан", "Адрес регистрации", "Место рождения")
    wide_min = {"Кем выдан": 80, "Адрес регистрации": 60, "Место рождения": 40}
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = columns[col_idx - 1] if col_idx <= len(columns) else ""
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(data) + 2)
        )
        if col_name in wide_cols:
            wmin = wide_min.get(col_name, 60)
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, wmin), 150)
        else:
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Рамки
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=1, max_row=len(data) + 1, min_col=1, max_col=len(columns)
    ):
        for cell in row:
            cell.border = thin_border

    output = Path(output_path)
    wb.save(output)
    return str(output)
